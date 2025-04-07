import random
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from flask import flash, redirect, url_for, current_app as app
from sqlalchemy import or_, and_, case, func, exists
from sqlalchemy.exc import OperationalError
from sqlalchemy.orm import joinedload
from functools import wraps

from collections import defaultdict
from . import db
from .models import Student, Session, Subject, Result, StudentTermSummary, class_teacher, Classes, Teacher, TermEnum, StudentClassHistory, FeePayment, class_subject
from application.auth.forms import SubjectResultForm

def require_session_and_term(f):
    """Decorator to ensure a current session and term are available."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        current_session, current_term = Session.get_current_session_and_term(include_term=True)
        if not current_session or not current_term:
            flash("No current session or term available.", "alert-danger")
            return redirect(url_for("admin.manage_sessions"))
        return f(current_session, current_term, *args, **kwargs)
    return decorated_function

def handle_db_errors(f):
    """Decorator to handle database errors consistently."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        try:
            return f(*args, **kwargs)
        except OperationalError as e:
            app.logger.error(f"Database error: {str(e)}")
            flash("A database error occurred. Please try again later.", "alert-danger")
            return redirect(url_for("admin.admin_dashboard"))
    return decorated_function

def validate_and_flash(form, redirect_url):
    """Validate a form and flash errors if validation fails."""
    if form.validate_on_submit():
        return True
    flash(f"Form validation failed: {form.errors}", "alert-danger")
    return redirect(redirect_url)

def generate_employee_id(first_name, last_name):
    """Generate a unique employee ID in the format first_name.last_name with number if needed."""
    # Convert to lowercase and remove any spaces or special characters
    base_id = f"{first_name.lower()}.{last_name.lower()}"

    try:
        # Query for existing teachers with similar employee IDs
        existing_teachers = Teacher.query.filter(
            Teacher.employee_id.like(f"{base_id}%")
        ).order_by(Teacher.employee_id).all()

        if not existing_teachers:
            return base_id  # Return base ID if no existing matches

        # Get all existing IDs
        existing_ids = [teacher.employee_id for teacher in existing_teachers]

        if base_id not in existing_ids:
            return base_id  # Return base ID if it's not taken

        # Find the next available number
        counter = 1
        while True:
            new_id = f"{base_id}{counter}"
            if new_id not in existing_ids:
                return new_id
            counter += 1

    except Exception as e:
        app.logger.error(f"Error generating employee ID: {str(e)}")
        raise
def get_teacher_classes(teacher_id, session_id, term):
    """Get classes assigned to a teacher for a specific session and term, falling back to previous if none exist."""
    current_assignments = db.session.query(Classes).join(class_teacher).filter(
        class_teacher.c.teacher_id == teacher_id,
        class_teacher.c.session_id == session_id,
        class_teacher.c.term == term
    ).all()

    if current_assignments:
        return current_assignments

    # Fallback to most recent previous assignment
    previous_assignments = db.session.query(Classes).join(class_teacher).filter(
        class_teacher.c.teacher_id == teacher_id,
        class_teacher.c.session_id < session_id
    ).order_by(class_teacher.c.session_id.desc(), class_teacher.c.term.desc()).all()

    return previous_assignments or []

def get_subjects_by_class_name(class_id, include_deactivated=False):
    """Get subjects assigned to a class via class_subject association."""
    try:
        # Query subjects linked to this class via class_subject
        query = Subject.query.join(class_subject).filter(class_subject.c.class_id == class_id)
        if not include_deactivated:
            query = query.filter(Subject.deactivated == False)
        subjects = query.order_by(Subject.name.asc()).all()
        return subjects
    except OperationalError as e:
        app.logger.error(f"Database error fetching subjects : {str(e)}")
        return []

def populate_form_with_results(form, subjects, results_dict):
    for subject in subjects:
        result = results_dict.get(subject.id)
        subject_form = SubjectResultForm(
            subject_id=subject.id,
            class_assessment='' if not result or result.class_assessment in [None, 0] else result.class_assessment,
            summative_test='' if not result or result.summative_test in [None, 0] else result.summative_test,
            exam='' if not result or result.exam in [None, 0] else result.exam,
            total='' if not result or result.total in [None, 0] else result.total,
            grade='' if not result or result.grade in [None, ''] else result.grade,
            remark='' if not result or result.remark in [None, ''] else result.remark,
        )
        form.subjects.append_entry(subject_form)

def calculate_grade(total):
    """Calculate grade based on total score."""
    if total is None: return ""
    if total >= 95: return "A+"
    elif total >= 80: return "A"
    elif total >= 70: return "B+"
    elif total >= 65: return "B"
    elif total >= 60: return "C+"
    elif total >= 50: return "C"
    elif total >= 40: return "D"
    elif total >= 30: return "E"
    return "F"

def generate_remark(total):
    """Generate remark based on total score."""
    if total is None: return ""
    if total >= 95: return "Outstanding"
    elif total >= 80: return "Excellent"
    elif total >= 70: return "Very Good"
    elif total >= 65: return "Good"
    elif total >= 60: return "Credit"
    elif total >= 50: return "Credit"
    elif total >= 40: return "Poor"
    elif total >= 30: return "Very Poor"
    return "Failed"

def get_threshold(average):
    """Determine the threshold range for a given average."""
    if average is None:
        return None
    thresholds = [0, 30, 40, 50, 60, 65, 70, 80, 90, 95]
    for i in range(len(thresholds) - 1):
        if thresholds[i] <= average < thresholds[i + 1]:
            return thresholds[i]
    return thresholds[-1] if average >= thresholds[-1] else thresholds[0]

def generate_principal_remark(average, student=None):
    """Generate principal's remark based on average, optionally using student object."""
    if average is None: return ""
    first_name = student.first_name if student else "Student"
    last_name = student.last_name if student else ""
    full_name = f"{first_name} {last_name}".strip()

    remarks = {
        95: [
            f"Truly exceptional, {full_name}! Your stellar achievement is a shining example for all students at this school.",
            f"Phenomenal results, {first_name}! You’ve set a benchmark for excellence across nursery, primary, and secondary levels.",
            f"Superb performance, {full_name}! You’ve outdone yourself and brought pride to our institution.",
            f"{first_name}, your outstanding dedication has paid off brilliantly! A top-tier result worthy of celebration.",
            f"Absolutely remarkable, {full_name}! Your consistency and brilliance are unmatched in this term.",
            f"{first_name} {last_name}, you are a star pupil! Your exceptional scores reflect your hard work and potential.",
            f"An extraordinary feat, {first_name}! Your academic prowess inspires both peers and teachers alike."
        ],
        90: [
            f"Exceptional performance, {full_name}! You are an inspiration to others in this school, keep it up!",
            f"Outstanding results, {first_name}! You’ve set high standards that others should aspire to achieve.",
            f"Excellent work, {full_name}! You’ve proven yourself as a top performer across all levels.",
            f"{first_name}, your brilliant effort this term is commendable! A fantastic example for your classmates.",
            f"Well done, {full_name}! Your dedication to excellence shines through in these impressive results.",
            f"{first_name} {last_name}, you’ve shown remarkable skill and focus! A truly outstanding achievement.",
            f"A splendid performance, {first_name}! Your consistency is paving the way for a bright future."
        ],
        80: [
            f"Very impressive performance, {full_name}! Aim for even higher success in the next term.",
            f"Great job, {first_name}! Your efforts are commendable and show promise for greater heights.",
            f"Consistently good results, {full_name}. Keep striving for excellence in all you do!",
            f"{first_name}, you’ve done wonderfully this term! Push a little more to reach the top.",
            f"Well done, {full_name}! Your hard work is evident, and greater achievements are within reach.",
            f"{first_name} {last_name}, a solid effort! Maintain this pace and aim for perfection.",
            f"Bravo, {first_name}! Your performance is strong—keep building on this excellent foundation."
        ],
        70: [
            f"Good performance, {full_name}! With a little more effort, you’ll excel further in your studies.",
            f"{first_name}, you’ve done well this term, but there’s room to improve and shine brighter.",
            f"Nice results, {full_name}! Strive for greater achievements with focus and determination.",
            f"{first_name} {last_name}, a commendable effort! Keep pushing to unlock your full potential.",
            f"Well done, {full_name}! You’re on a good path—add more diligence for outstanding success.",
            f"{first_name}, your performance is solid! Aim higher by working on your weaker areas.",
            f"A promising result, {full_name}! Consistency and extra effort will take you far."
        ],
        65: [
            f"Satisfactory effort, {full_name}. Aim to push beyond the basics for better outcomes.",
            f"{first_name}, you’re doing alright, but consistency is key to reaching higher grades.",
            f"Good start, {full_name}! Work harder to climb to the next level of success.",
            f"{first_name} {last_name}, a fair performance! More focus will help you improve greatly.",
            f"{full_name}, you’ve laid a foundation—build on it with dedication and study.",
            f"{first_name}, your effort is noted! Aim higher by strengthening your skills this term.",
            f"Acceptable results, {full_name}. Step up your game to achieve more in your academics."
        ],
        60: [
            f"Fair performance, {full_name}. Keep pushing for better results in the coming terms.",
            f"{first_name}, you’re on the right track—focus on improving your weaker subjects.",
            f"A decent effort, {full_name}! Keep working hard to achieve more and excel.",
            f"{first_name} {last_name}, you’ve shown some progress! More effort will yield greater rewards.",
            f"{full_name}, an okay result—strengthen your study habits for a stronger outcome.",
            f"{first_name}, you’re moving forward! Target your challenges to boost your scores.",
            f"Moderate success, {full_name}. Determination will help you rise above this level."
        ],
        50: [
            f"An average performance, {full_name}. Aim for consistent improvement in all subjects.",
            f"{first_name}, your results are satisfactory, but you can do much better with effort.",
            f"A fair effort, {full_name}! Focus and determination will lead to greater success.",
            f"{first_name} {last_name}, you’re at the midpoint—work harder to stand out next term.",
            f"{full_name}, this is a starting point! Build your skills to improve your grades.",
            f"{first_name}, an acceptable result—push beyond average with serious study habits.",
            f"Room for growth, {full_name}! Take your studies seriously to see better outcomes."
        ],
        40: [
            f"Below average, {full_name}. Significant improvement is required to succeed.",
            f"{first_name}, your performance is concerning—seek guidance to improve this term.",
            f"You need to focus more, {full_name}! Extra effort will help you catch up.",
            f"{first_name} {last_name}, this isn’t your best—work with teachers to do better.",
            f"{full_name}, your results need attention! Take action to boost your understanding.",
            f"{first_name}, a weak performance—commit to studying harder for progress.",
            f"Low scores, {full_name}. Let’s work together to turn this around next term."
        ],
        30: [
            f"Poor results, {full_name}. A lot more effort is needed to improve your standing.",
            f"{first_name}, your performance requires immediate attention and serious work.",
            f"Seek extra help, {full_name}! Address your challenges to succeed this year.",
            f"{first_name} {last_name}, this is below expectations—let’s find ways to improve.",
            f"{full_name}, your scores are low—dedicate more time to your books and lessons.",
            f"{first_name}, a tough term—extra support can help you overcome these struggles.",
            f"Very concerning, {full_name}! Act now to avoid falling further behind."
        ],
        0: [
            f"Unacceptable performance, {full_name}. Take your studies seriously from now on.",
            f"{first_name}, drastic improvement is needed to progress in your academics.",
            f"This is not satisfactory, {full_name}! Immediate action is required to improve.",
            f"{first_name} {last_name}, a worrying result—commit to change this outcome.",
            f"{full_name}, your effort is lacking—focus on your education starting today.",
            f"{first_name}, this performance is poor—let’s work together to fix this urgently.",
            f"No progress, {full_name}! Serious dedication is needed to move forward."
        ]
    }
    for threshold, options in sorted(remarks.items(), reverse=True):
        if average >= threshold:
            return random.choice(options)
    return ""

def generate_teacher_remark(average, student=None):
    """Generate teacher's remark based on average, optionally using student object."""
    if average is None: return ""
    first_name = student.first_name if student else "Student"
    last_name = student.last_name if student else ""
    full_name = f"{first_name} {last_name}".strip()

    remarks = {
        95: [
            f"Absolutely outstanding, {full_name}! You’re a model student for others to follow.",
            f"Exceptional effort, {first_name}! Your results shine brightly this term—keep it up!",
            f"Top-tier performance, {full_name}! Your hard work is truly impressive and inspiring.",
            f"{first_name} {last_name}, you’ve excelled brilliantly! A star in every subject.",
            f"{full_name}, your dedication is remarkable! You’ve set a high standard this term.",
            f"{first_name}, an amazing result! Your focus and brilliance are commendable.",
            f"Perfect scores, {full_name}! You’re a joy to teach and a pride to this class."
        ],
        90: [
            f"Outstanding performance, {full_name}! Keep shining in all your subjects.",
            f"{first_name}, your dedication to learning is inspiring—well done this term!",
            f"An exemplary effort, {full_name}! Your hard work has paid off wonderfully.",
            f"{first_name} {last_name}, fantastic results! You’re a leader among your peers.",
            f"{full_name}, you’ve done exceptionally well! Maintain this excellent momentum.",
            f"{first_name}, a brilliant term! Your consistency is paving your path to success.",
            f"Great work, {full_name}! Your effort stands out as one of the best this term."
        ],
        80: [
            f"A good performance, {full_name}! Focus on consistent excellence moving forward.",
            f"{first_name}, your hard work is paying off—keep aiming higher each term!",
            f"You’ve done very well, {full_name}! Push your limits for even better results.",
            f"{first_name} {last_name}, solid effort! You’re close to the top—keep going!",
            f"{full_name}, a strong showing! Build on this to reach outstanding heights.",
            f"{first_name}, well done! Your progress is notable—aim for perfection next time.",
            f"Impressive, {full_name}! Stay committed to rise even higher in your studies."
        ],
        70: [
            f"Good result, {full_name}! Practice more to perfect your skills this term.",
            f"{first_name}, you’re doing well—focus on weaker areas to improve further.",
            f"A good effort, {full_name}! Strive to achieve even better in your subjects.",
            f"{first_name} {last_name}, nice work! Extra study will boost your grades.",
            f"{full_name}, you’re on track—target your challenges for greater success.",
            f"{first_name}, a fair result! Keep up the effort to reach higher scores.",
            f"Promising, {full_name}! More attention to detail will lift your performance."
        ],
        65: [
            f"Decent result, {full_name}. Build on this foundation for better grades.",
            f"{first_name}, you’re improving—more effort will yield stronger outcomes.",
            f"Satisfactory, {full_name}! Aim higher with consistent practice and focus.",
            f"{first_name} {last_name}, a good base! Work harder to climb higher.",
            f"{full_name}, you’re getting there—strengthen your study habits now.",
            f"{first_name}, fair effort! Target key areas to improve your standing.",
            f"Acceptable, {full_name}! Step up your game to excel in your class."
        ],
        60: [
            f"Fair result, {full_name}. Put in more effort to rise above this level.",
            f"{first_name}, you’re progressing—focus on difficult topics to improve.",
            f"With more effort, {full_name}, your performance will improve significantly!",
            f"{first_name} {last_name}, an okay term! Push harder for better results.",
            f"{full_name}, you’re moving along—work on weak spots for growth.",
            f"{first_name}, a modest outcome! Extra study will make a big difference.",
            f"Moderate effort, {full_name}! Keep pushing to unlock your potential."
        ],
        50: [
            f"An average performance, {full_name}. Work on improving your fundamentals.",
            f"{first_name}, you’ve done okay—additional practice will help you grow.",
            f"Keep working on weaker subjects, {full_name}, to build a strong foundation!",
            f"{first_name} {last_name}, fair results! Aim higher with more effort.",
            f"{full_name}, this is average—focus more to improve your scores.",
            f"{first_name}, a starting point! Serious study will lift you up.",
            f"Room to grow, {full_name}! Dedicate time to see better progress."
        ],
        40: [
            f"Below average, {full_name}. Concentrate on improving your basics this term.",
            f"{first_name}, your performance is concerning—seek help to address challenges.",
            f"Weak results, {full_name}! Focus on understanding the core concepts.",
            f"{first_name} {last_name}, this needs work—let’s improve together.",
            f"{full_name}, low scores—commit to studying harder for better outcomes.",
            f"{first_name}, a tough term—extra support can turn this around.",
            f"Struggling, {full_name}! Let’s tackle your weak areas urgently."
        ],
        30: [
            f"Very poor performance, {full_name}. A focused study plan is necessary now.",
            f"{first_name}, you need additional support to overcome your difficulties.",
            f"Your results are below expectations, {full_name}! Put in much more effort.",
            f"{first_name} {last_name}, this is concerning—let’s work on improvement.",
            f"{full_name}, low grades—spend more time with your books and teachers.",
            f"{first_name}, a challenging term—seek help to boost your understanding.",
            f"Poor showing, {full_name}! Act quickly to improve your academics."
        ],
        0: [
            f"Unacceptable, {full_name}. Urgent attention to your studies is required.",
            f"{first_name}, a very poor performance—take your academics seriously now!",
            f"This is concerning, {full_name}! Take your studies more seriously.",
            f"{first_name} {last_name}, no progress—let’s change this urgently.",
            f"{full_name}, your effort is lacking—focus on your education today.",
            f"{first_name}, a worrying result—commit to improvement immediately.",
            f"Critical, {full_name}! Serious dedication is needed to move forward."
        ]
    }
    for threshold, options in sorted(remarks.items(), reverse=True):
        if average >= threshold:
            return random.choice(options)
    return ""

def get_last_term(current_term):
    """Get the previous term."""
    term_order = {TermEnum.FIRST.value: 0, TermEnum.SECOND.value: 1, TermEnum.THIRD.value: 2}
    current_idx = term_order.get(current_term)
    if current_idx is None or current_idx == 0:
        return None
    return list(term_order.keys())[current_idx - 1]

def calculate_average(results):
    """Calculate average score and subject count."""
    unique_results = {r.subject_id: r for r in results}.values()
    non_zero_results = [r.total for r in unique_results if r.total is not None and r.total > 0]
    total_sum = sum(non_zero_results) if non_zero_results else 0
    count = len(non_zero_results)
    return total_sum / count if count > 0 else 0, count

def calculate_cumulative_average(student_id, term, session_id):
    """Calculate cumulative average based on term averages within a session."""
    try:
        # Define term order using TermEnum
        term_order = {
            TermEnum.FIRST.value: 1,
            TermEnum.SECOND.value: 2,
            TermEnum.THIRD.value: 3
        }
        current_term_order = term_order.get(term, 1)

        # Query term averages from StudentTermSummary for the session
        term_summaries = StudentTermSummary.query.filter_by(
            student_id=student_id,
            session_id=session_id
        ).all()

        # Extract term averages up to the current term
        term_averages = {}
        for summary in term_summaries:
            if summary.term in term_order and summary.term_average is not None:
                term_averages[summary.term] = summary.term_average

        # Calculate cumulative average based on the current term
        if current_term_order == 1:  # First Term
            return term_averages.get(TermEnum.FIRST.value, 0)
        elif current_term_order == 2:  # Second Term
            first_avg = term_averages.get(TermEnum.FIRST.value, 0)
            second_avg = term_averages.get(TermEnum.SECOND.value, 0)
            valid_avgs = [avg for avg in [first_avg, second_avg] if avg > 0]
            return sum(valid_avgs) / len(valid_avgs) if valid_avgs else 0
        elif current_term_order == 3:  # Third Term
            first_avg = term_averages.get(TermEnum.FIRST.value, 0)
            second_avg = term_averages.get(TermEnum.SECOND.value, 0)
            third_avg = term_averages.get(TermEnum.THIRD.value, 0)
            valid_avgs = [avg for avg in [first_avg, second_avg, third_avg] if avg > 0]
            return sum(valid_avgs) / len(valid_avgs) if valid_avgs else 0
        return 0  # Default case if term is invalid
    except OperationalError as e:
        app.logger.error(f"Database error in calculate_cumulative_average: {str(e)}")
        raise

def update_results_helper(student, term, session_id, form, result_form, class_id):
    """Update student results with form data, using session_id."""
    try:
        next_term_begins = result_form.next_term_begins.data
        date_issued = result_form.date_issued.data
        position = result_form.position.data

        for subject_form in form.subjects:
            subject_id_raw = subject_form.subject_id.data
            try:
                subject_id = int(subject_id_raw)
            except (ValueError, TypeError) as e:
                app.logger.error(f"Invalid subject_id '{subject_id_raw}' for student {student.id}: {str(e)}")
                raise ValueError(f"Invalid subject_id: {subject_id_raw}")

            data = {
                "class_assessment": subject_form.class_assessment.data or None,
                "summative_test": subject_form.summative_test.data or None,
                "exam": subject_form.exam.data or None,
                "next_term_begins": next_term_begins or None,
                "date_issued": date_issued or None,
                "position": position or None,
            }
            save_result(student.id, subject_id, term, session_id, data, class_id=class_id)
    except Exception as e:
        app.logger.error(f"Error updating results for student {student.id}: {str(e)}")
        raise

def save_result(student_id, subject_id, term, session_id, data, class_id):
    try:
        class_assessment = int(data.get("class_assessment", 0)) if data.get("class_assessment") not in [None, '', '0'] else None
        summative_test = int(data.get("summative_test", 0)) if data.get("summative_test") not in [None, '', '0'] else None
        exam = int(data.get("exam", 0)) if data.get("exam") not in [None, '', '0'] else None
        position = data.get("position")
        next_term_begins = data.get("next_term_begins")
        date_issued = data.get("date_issued")

        student = Student.query.filter_by(id=student_id).first()
        if not student:
            raise ValueError(f"Student with ID {student_id} not found.")

        # Calculate current_average BEFORE applying the new result
        existing_results = Result.query.filter_by(
            student_id=student_id,
            term=term,
            session_id=session_id,
            class_id=class_id
        ).all()
        current_average = round(calculate_average(existing_results)[0], 1) if existing_results else None
        current_threshold = get_threshold(current_average) if current_average is not None else None

        result = None
        if subject_id:
            total_components = [x for x in [class_assessment, summative_test, exam] if x is not None]
            total = sum(total_components) if total_components else None
            grade = calculate_grade(total) if total is not None else ''
            remark = generate_remark(total) if total is not None else ''

            result = Result.query.filter_by(
                student_id=student_id,
                subject_id=subject_id,
                term=term,
                session_id=session_id,
                class_id=class_id
            ).first()

            if result or any([class_assessment, summative_test, exam]):
                if result:
                    # Update existing result
                    result.class_assessment = class_assessment
                    result.summative_test = summative_test
                    result.exam = exam
                    result.total = total
                    result.grade = grade
                    result.remark = remark
                else:
                    # Create new result
                    result = Result(
                        student_id=student_id,
                        subject_id=subject_id,
                        term=term,
                        session_id=session_id,
                        class_id=class_id,
                        class_assessment=class_assessment,
                        summative_test=summative_test,
                        exam=exam,
                        total=total,
                        grade=grade,
                        remark=remark,
                    )
                    db.session.add(result)

                db.session.flush()

        # Calculate new aggregates AFTER applying the result
        grand_total, average, cumulative_average, last_term_average, subjects_offered = calculate_results(student_id, term, session_id, class_id)
        new_threshold = get_threshold(average) if average is not None else None

        term_summary = StudentTermSummary.query.filter_by(
            student_id=student_id,
            term=term,
            session_id=session_id,
            class_id=class_id
        ).first()

        if term_summary:
            term_summary.grand_total = grand_total
            term_summary.term_average = average
            term_summary.cumulative_average = cumulative_average
            term_summary.last_term_average = last_term_average
            term_summary.subjects_offered = subjects_offered

            if average is not None:
                if current_threshold is None or current_threshold != new_threshold or term_summary.principal_remark is None:
                    term_summary.principal_remark = generate_principal_remark(average, student)
                    term_summary.teacher_remark = generate_teacher_remark(average, student)

            if position is not None:
                term_summary.position = None if position == "" else position
            if next_term_begins is not None:
                term_summary.next_term_begins = None if next_term_begins == "" else next_term_begins
            if date_issued is not None:
                term_summary.date_issued = None if date_issued == "" else date_issued
        else:
            term_summary = StudentTermSummary(
                student_id=student_id,
                term=term,
                session_id=session_id,
                class_id=class_id,
                grand_total=grand_total,
                term_average=average,
                cumulative_average=cumulative_average,
                last_term_average=last_term_average,
                subjects_offered=subjects_offered,
                principal_remark=generate_principal_remark(average, student) if average is not None else None,
                teacher_remark=generate_teacher_remark(average, student) if average is not None else None,
                position=None if position == "" else position,
                next_term_begins=None if next_term_begins == "" else next_term_begins,
                date_issued=None if date_issued == "" else date_issued,
            )
            db.session.add(term_summary)

        db.session.commit()
        return result
    except ValueError as e:
        db.session.rollback()
        app.logger.error(f"Invalid input in save_result: {str(e)}")
        raise
    except OperationalError as e:
        db.session.rollback()
        app.logger.error(f"Database error in save_result: {str(e)}")
        raise

def save_class_wide_fields(class_id, session_id, term, data):
    """Save class-wide fields (next_term_begins, date_issued) for all active students in a class."""
    try:
        next_term_begins = data.get("next_term_begins")
        date_issued = data.get("date_issued")
        app.logger.info(f"Received next_term_begins: {next_term_begins}, date_issued: {date_issued}")

        # Fetch all active students in the class for this session
        student_class_histories = StudentClassHistory.query.filter_by(
            session_id=session_id,
            class_id=class_id,
            is_active=True,
            leave_date=None
        ).all()
        student_ids = [history.student_id for history in student_class_histories]

        # Update existing StudentTermSummary records
        term_summaries = StudentTermSummary.query.filter(
            StudentTermSummary.student_id.in_(student_ids),
            StudentTermSummary.term == term,
            StudentTermSummary.session_id == session_id,
            StudentTermSummary.class_id == class_id
        ).all()

        updated = False
        for term_summary in term_summaries:
            app.logger.info(f"Before update - term_summary.next_term_begins: {term_summary.next_term_begins}, date_issued: {term_summary.date_issued}")
            if next_term_begins is not None:  # None means no update from client
                if next_term_begins == "":
                    term_summary.next_term_begins = None
                    app.logger.info("next_term_begins cleared to None")
                else:
                    term_summary.next_term_begins = next_term_begins
                    app.logger.info(f"next_term_begins set to: {next_term_begins}")
                updated = True
            if date_issued is not None:  # None means no update from client
                if date_issued == "":
                    term_summary.date_issued = None
                    app.logger.info("date_issued cleared to None")
                else:
                    term_summary.date_issued = date_issued
                    app.logger.info(f"date_issued set to: {date_issued}")
                updated = True
            app.logger.info(f"After update - term_summary.next_term_begins: {term_summary.next_term_begins}, date_issued: {term_summary.date_issued}")

        db.session.flush()  # Ensure changes are staged
        return len(term_summaries) if updated else 0  # Return count only if changes were made
    except OperationalError as e:
        db.session.rollback()
        app.logger.error(f"Database error in save_class_wide_fields: {str(e)}")
        raise

def calculate_results(student_id, term, session_id, class_id):
    """Calculate student results and averages with optimized queries."""
    try:
        # Filter results by class_id and session_id for the current term
        results = Result.query.filter_by(
            student_id=student_id,
            term=term,
            session_id=session_id,
            class_id=class_id
        ).all()
        grand_total = sum(r.total for r in results if r.total is not None) or 0
        average, subjects_offered = calculate_average(results)
        average = round(average, 1) if average else 0

        # Calculate cumulative average based on term averages
        cumulative_average = round(calculate_cumulative_average(student_id, term, session_id), 1)

        # Calculate last term average
        last_term = get_last_term(term)
        last_term_results = Result.query.filter_by(
            student_id=student_id,
            term=last_term,
            session_id=session_id,
            class_id=class_id
        ).all() if last_term else []
        last_term_average = round(calculate_average(last_term_results)[0], 1) if last_term_results else None

        return grand_total, average, cumulative_average, last_term_average, subjects_offered
    except OperationalError as e:
        app.logger.error(f"Database error in calculate_results: {str(e)}")
        raise

def prepare_broadsheet_data(students, subjects, term, session_year, class_id, session_id):
    """Prepare data for broadsheet with class_id and session_id."""
    broadsheet_data = []
    subject_averages = {subject.id: {"total": 0, "count": 0} for subject in subjects}
    try:
        for student in students:
            student_results = {
                "student": student,
                "results": {subject.id: None for subject in subjects},
                "grand_total": "",
                "average": "",
                "cumulative_average": "",
                "position": None,
                "principal_remark": "",
                "teacher_remark": "",
                "next_term_begins": None,
                "date_issued": None
            }

            results = Result.query.filter_by(
                student_id=student.id,
                term=term,
                session_id=session_id,
                class_id=class_id
            ).all()

            term_summary = StudentTermSummary.query.filter_by(
                student_id=student.id,
                term=term,
                session_id=session_id,
                class_id=class_id
            ).first()

            for result in results:
                if result.subject_id in student_results["results"]:
                    student_results["results"][result.subject_id] = result
                    if result.total is not None and result.total > 0:
                        subject_averages[result.subject_id]["total"] += result.total
                        subject_averages[result.subject_id]["count"] += 1

            if term_summary:
                student_results["grand_total"] = int(term_summary.grand_total) if term_summary.grand_total not in ("", None) else ""
                student_results["average"] = float(term_summary.term_average) if term_summary.term_average not in ("", None) else ""
                student_results["cumulative_average"] = float(term_summary.cumulative_average) if term_summary.cumulative_average not in ("", None) else ""
                student_results["position"] = term_summary.position if term_summary.position not in ("", None) else ""
                student_results["principal_remark"] = term_summary.principal_remark if term_summary.principal_remark not in ("", None) else ""
                student_results["teacher_remark"] = term_summary.teacher_remark if term_summary.teacher_remark not in ("", None) else ""
                student_results["next_term_begins"] = term_summary.next_term_begins
                student_results["date_issued"] = term_summary.date_issued

            broadsheet_data.append(student_results)

        for subject_id, values in subject_averages.items():
            values["average"] = round(values["total"] / values["count"], 1) if values["count"] else ""

        broadsheet_data.sort(
            key=lambda x: float(x["average"]) if isinstance(x["average"], (int, float, str)) and x["average"] not in ("", None) else 0,
            reverse=True
        )

        return broadsheet_data, subject_averages
    except OperationalError as e:
        app.logger.error(f"Database error preparing broadsheet data: {str(e)}")
        raise

def generate_excel_broadsheet(class_name, term, session_year, broadsheet_data, subjects, subject_averages):
    """Generate an Excel broadsheet."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Broadsheet_{class_name}_{term}"[:31]

    header_font = Font(bold=True, size=16, name="Times New Roman")
    sub_header_font = Font(bold=True, size=12, name="Times New Roman")
    cell_font = Font(bold=True, size=12, name="Times New Roman")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    alignment = Alignment(horizontal="center", vertical="center")

    # Calculate the last column based on the number of students and their columns
    last_column = 2 + len(broadsheet_data) * 4  # 4 columns per student (C/A, S/T, Exam, Total)
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_column)
    sheet["B1"] = f"Broadsheet for {class_name} - Term: {term}, Session: {session_year}"
    sheet["B1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet["B1"].font = header_font

    # Set column widths
    sheet.column_dimensions[get_column_letter(1)].width = 30
    sheet.column_dimensions[get_column_letter(last_column)].width = 15

    # Create headers with student names
    headers = [""]
    for student_data in broadsheet_data:
        student = student_data["student"]
        headers.extend([f"{student.first_name} {student.last_name}", "", "", ""])
    headers.append("Class Average")
    sheet.append(headers)

    # Merge cells for student names and apply styling
    for i, student_data in enumerate(broadsheet_data, start=2):
        start_col = 2 + (i - 2) * 4
        end_col = start_col + 3
        for col in range(start_col, end_col + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 7
        sheet.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        cell = sheet.cell(row=2, column=start_col)
        cell.alignment = alignment
        cell.font = sub_header_font

    # Add sub-headers for C/A, S/T, Exam, Total
    sub_headers = ["Subjects"] + ["C/A", "S/T", "Exam", "Total"] * len(broadsheet_data) + [""]
    sheet.append(sub_headers)

    # Add subject rows with student results
    for subject in subjects:
        row = [subject.name]
        for student_data in broadsheet_data:
            result = student_data["results"][subject.id]
            row.extend([
                result.class_assessment if result and result.class_assessment is not None else "",
                result.summative_test if result and result.summative_test is not None else "",
                result.exam if result and result.exam is not None else "",
                result.total if result and result.total is not None else "",
            ])
        row.append(subject_averages[subject.id]["average"] or "")
        sheet.append(row)

    # Add aggregate rows: Grand Total, Average, Position
    sheet.append([""])  # Empty row for spacing
    sheet.append(["Grand Total"] + sum([[ "", "", "", d["grand_total"] or ""] for d in broadsheet_data], []) + [""])
    sheet.append(["Average"] + sum([[ "", "", "", d["average"] or ""] for d in broadsheet_data], []) + [""])
    sheet.append(["Cumulative Average"] + sum([[ "", "", "", d["cumulative_average"] or ""] for d in broadsheet_data], []) + [""])
    sheet.append(["Position"] + sum([[ "", "", "", d["position"] or ""] for d in broadsheet_data], []) + [""])

    # Set page setup for printing
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_margins.top = 0.75
    sheet.page_margins.bottom = 0.75
    sheet.page_margins.left = 0.252
    sheet.page_margins.right = 0.252
    sheet.page_margins.header = 0.299
    sheet.page_margins.footer = 0.299

    # Apply borders, fonts, and alignment to all cells
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = border
            cell.font = cell_font
            cell.alignment = Alignment(horizontal="left" if cell.column == 1 else "center", vertical="center")

    # Save the workbook to a BytesIO stream
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output

def group_students_by_class(students):
    """Group students by class name, sorted by hierarchy and student name."""
    if not students:
        app.logger.debug("No students provided to group_students_by_class")
        return {}

    students_classes = defaultdict(list)
    class_hierarchies = {}
    for item in students:
        try:
            student, class_name, hierarchy = item
            if not isinstance(student, Student) or not class_name:
                app.logger.warning(f"Invalid student or class_name: {student}, {class_name}")
                continue
            students_classes[class_name].append(student)
            class_hierarchies[class_name] = hierarchy  # Store hierarchy directly from query
        except (ValueError, TypeError) as e:
            app.logger.warning(f"Failed to unpack student tuple: {item}, error: {str(e)}")
            continue

    # Sort students within each class by name
    for class_name in students_classes:
        students_classes[class_name].sort(key=lambda s: (s.first_name.lower(), s.last_name.lower()))

    # Sort classes by hierarchy
    sorted_classes = sorted(
        students_classes.items(),
        key=lambda item: class_hierarchies.get(item[0], float('inf'))
    )
    return dict(sorted_classes)


def get_students_query(current_session, term=None):
    term_order = {t.value: i for i, t in enumerate(TermEnum, 1)}
    current_term = Session.get_current_session_and_term(include_term=True)[1]
    current_term_order = term_order.get(current_term.value, 1) if current_term else 1
    target_term = term if term else current_term.value
    target_term_order = term_order.get(target_term, current_term_order)

    # Base query for latest enrollment
    latest_enrollment_subquery = (
        db.session.query(
            StudentClassHistory.student_id,
            func.max(StudentClassHistory.join_date).label('latest_join')
        )
        .filter(StudentClassHistory.session_id == current_session.id)
        .group_by(StudentClassHistory.student_id)
        .subquery('latest_enrollment')
    )

    query = (
        db.session.query(Student, Classes.name.label('class_name'), Classes.hierarchy)
        .join(
            latest_enrollment_subquery,
            Student.id == latest_enrollment_subquery.c.student_id,
            isouter=True
        )
        .join(
            StudentClassHistory,
            and_(
                Student.id == StudentClassHistory.student_id,
                StudentClassHistory.session_id == current_session.id,
                StudentClassHistory.join_date == latest_enrollment_subquery.c.latest_join,
                StudentClassHistory.is_active == True,
                StudentClassHistory.leave_date.is_(None),
                case(
                    *[(StudentClassHistory.start_term == t.value, i) for i, t in enumerate(TermEnum, 1)],
                    else_=1
                ) <= target_term_order,
                or_(
                    StudentClassHistory.end_term.is_(None),
                    case(
                        *[(StudentClassHistory.end_term == t.value, i) for i, t in enumerate(TermEnum, 1)],
                        else_=4
                    ) >= target_term_order
                )
            ),
            isouter=True
        )
        .join(Classes, StudentClassHistory.class_id == Classes.id, isouter=True)
    )

    return query.options(joinedload(Student.fee_payments), joinedload(Student.class_history))

def apply_filters_to_students_query(students_query, enrollment_status, fee_status, approval_status, current_session, current_term, term=None):
    """Apply filters considering term-specific enrollment and promotion status."""
    term_order = {t.value: i for i, t in enumerate(TermEnum, 1)}
    current_term_order = term_order.get(current_term.value, 1)
    target_term = term if term else current_term.value
    target_term_order = term_order.get(target_term, current_term_order)

    # Define session range for inactivity (2023/2024 onwards)
    base_session = Session.query.filter_by(year="2023/2024").first()
    base_session_id = base_session.id if base_session else current_session.id  # Fallback

    if enrollment_status == 'active':
        students_query = students_query.filter(
            StudentClassHistory.is_active == True,
            StudentClassHistory.leave_date.is_(None),
            StudentClassHistory.session_id == current_session.id,
            case(
                *[(StudentClassHistory.start_term == t.value, i) for i, t in enumerate(TermEnum, 1)],
                else_=1
            ) <= target_term_order,
            or_(
                StudentClassHistory.end_term.is_(None),
                case(
                    *[(StudentClassHistory.end_term == t.value, i) for i, t in enumerate(TermEnum, 1)],
                    else_=4
                ) >= target_term_order
            )
        )
    elif enrollment_status == 'inactive':
        # Precompute the count of students in the current session
        session_student_count = db.session.query(
            func.count(StudentClassHistory.id)
        ).filter(
            StudentClassHistory.session_id == current_session.id
        ).scalar() or 0

        # Subquery for students with no active records in the current session and term
        no_record_subquery = (
            ~exists().where(
                and_(
                    StudentClassHistory.student_id == Student.id,
                    StudentClassHistory.session_id == current_session.id,
                    StudentClassHistory.is_active == True,
                    StudentClassHistory.leave_date.is_(None),
                    case(
                        *[(StudentClassHistory.start_term == t.value, i) for i, t in enumerate(TermEnum, 1)],
                        else_=1
                    ) <= target_term_order,
                    or_(
                        StudentClassHistory.end_term.is_(None),
                        case(
                            *[(StudentClassHistory.end_term == t.value, i) for i, t in enumerate(TermEnum, 1)],
                            else_=4
                        ) >= target_term_order
                    )
                )
            ).correlate(Student)
        )

        students_query = (
            db.session.query(Student, Classes.name.label('class_name'), Classes.hierarchy)
            .join(StudentClassHistory, StudentClassHistory.student_id == Student.id, isouter=True)
            .join(Classes, StudentClassHistory.class_id == Classes.id, isouter=True)
            .filter(
                or_(
                    and_(
                        StudentClassHistory.session_id >= base_session_id,
                        or_(
                            StudentClassHistory.is_active == False,
                            and_(
                                StudentClassHistory.leave_date.isnot(None),
                                case(
                                    *[(StudentClassHistory.end_term == t.value, i) for i, t in enumerate(TermEnum, 1)],
                                    else_=0
                                ) <= target_term_order
                            )
                        )
                    ),
                    and_(
                        no_record_subquery,
                        session_student_count > 10
                    )
                )
            )
        )

    if enrollment_status != 'inactive' and fee_status:
        students_query = students_query.outerjoin(
            FeePayment,
            and_(
                FeePayment.student_id == Student.id,
                FeePayment.session_id == current_session.id,
                FeePayment.term == target_term
            )
        )
        if fee_status == 'paid':
            students_query = students_query.filter(FeePayment.has_paid_fee == True)
        elif fee_status == 'unpaid':
            students_query = students_query.filter(
                or_(
                    FeePayment.has_paid_fee == False,
                    FeePayment.id.is_(None)
                )
            )

    if approval_status:
        students_query = students_query.filter(
            Student.approved == (approval_status == 'approved')
        )

    return students_query

