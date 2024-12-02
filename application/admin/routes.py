"""
Generic functions/routes of the Admin Dashboard

    admin_dashboard: Displays the admin dashboard
    manage_sessions: Manages the current session
    change_session: Changes the current session
"""

"""
Route for managing sessions.

This route allows administrators to manage sessions. Only users with admin privileges can access this route.
Administrators can select a session from a form and change the current session to the selected one.

Returns:
    If the form is submitted successfully, the user is redirected to the "change_session" route with the selected session ID.
    Otherwise, the user is rendered the "admin/manage_sessions.html" template with the current session and the session form.
"""
from datetime import datetime
from collections import defaultdict
from io import BytesIO
from urllib.parse import unquote
import openpyxl
from sqlalchemy.orm import joinedload
from flask_wtf.csrf import CSRFError
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from flask_login import login_required, current_user
from flask import (
    abort,
    render_template,
    redirect,
    url_for,
    flash,
    request,
    Response,
    # make_response,
    current_app as app,
)

from . import admin_bp
from ..helpers import (
    get_subjects_by_class_name,
    update_results,
    calculate_results,
    db,
    random,
    string,
    populate_form_with_results,
)
from ..models import Student, User, Subject, Result, Session, StudentClassHistory

from ..auth.forms import (
    EditStudentForm,
    ResultForm,
    SubjectForm,
    DeleteForm,
    SessionForm,
    ApproveForm,
    classForm,
    ManageResultsForm,
)


@admin_bp.errorhandler(CSRFError)
def handle_csrf_error(e):
    app.logger.warning(f"CSRF error: {e.description}")
    flash("The form submission has expired. Please try again.", "alert alert-danger")
    return redirect(url_for("admins.admin_dashboard"))


@admin_bp.errorhandler(500)
def internal_error(error):
    app.logger.error(f"Server Error: {error}")
    return redirect(url_for("admins.admin_dashboard"))


@admin_bp.errorhandler(404)
def not_found_error(error):
    app.logger.warning(f"404 Error: {error}")
    return redirect(url_for("admins.admin_dashboard"))


@admin_bp.before_request
@login_required
def admin_before_request():
    if not current_user.is_admin:
        flash("You are not authorized to access this page.", "alert alert-danger")
        app.logger.warning(
            f"Unauthorized access attempt by user {current_user.username}"
        )
        return redirect(url_for("main.index"))


@admin_bp.route("/dashboard")
def admin_dashboard():
    app.logger.info(f"Admin dashboard accessed by user {current_user.username}")
    return render_template("admin/index.html")


@admin_bp.route("/manage_sessions", methods=["GET", "POST"])
@login_required
def manage_sessions():
    if not current_user.is_admin:
        abort(403)  # Restrict access for non-admins

    form = SessionForm()

    # Fetch all sessions and populate choices
    sessions = Session.query.all()
    form.session.choices = [(session.id, session.year) for session in sessions]

    # Include term options
    terms = ["First Term", "Second Term", "Third Term"]
    form.term.choices = [(term, term) for term in terms]
    
    # Retrieve current session and term
    # current_session, current_term = Session.get_current_session_and_term(include_term=True)

    if form.validate_on_submit():
        session_id = form.session.data
        term = form.term.data

        # Update the session and term in the database
        updated_session = Session.set_current_session(session_id, term)
        
        if updated_session:
            flash(
                f"Current session set to {updated_session.year} ({term}).",
                "alert alert-success",
            )
        else:
            flash("Failed to update the current session and term.", "alert alert-danger")

        return redirect(url_for("admins.manage_sessions"))
        
    # # Retrieve current session and term
    current_session, current_term = Session.get_current_session_and_term(include_term=True)
    
    return render_template(
        "admin/manage_sessions.html",
        form=form,
        current_session=current_session,
        current_term=current_term,
    )


""""
Student Management Section

This section displays the basic view of the student management page
The page displays the list of students and their details
It also provides links to edit, delete, approve, and deactivate students

    approve_students: Displays the list of students and their approval status
    approve_student: Approves a student
    deactivate_student: Deactivates a student
    regenerate_password: Regenerates the password for a student
    manage_classes: Displays the list of classes
    view_class_by_session: Displays the list of students in a class for a given session
    students_by_class: Displays the list of students in a class
    promote_student: Promotes a student to the next class
    edit_student: Edits the details of a student
    delete_student: Deletes a student
"""

@admin_bp.route("/admin/approve_students", methods=["GET", "POST"])
@login_required
def approve_students():
    """
    View function for approving students in the admin panel.

    This function renders the approve students page, where admins can:
    - Retrieve students and their class history for the current session and term.
    - Group the students by their class name for approval actions.
    - Handle approval, deactivation, and regeneration actions for students.

    Returns:
        The rendered template for the approve students page.

    Raises:
        403: If the current user is not an admin.
    """
    if not current_user.is_admin:
        abort(403)  # Restrict access for non-admins

    # Get current session and current term directly from the Session model
    current_session, current_term = Session.get_current_session_and_term(include_term=True)

    # Fetch the current session ID using the current session year
    selected_session_id = Session.query.filter_by(year=current_session).first().id

    # Fetch students and their class history for the current session
    students = (
        db.session.query(Student, StudentClassHistory.class_name)
        .join(
            StudentClassHistory,
            (Student.id == StudentClassHistory.student_id)
            & (StudentClassHistory.session_id == selected_session_id),
        )
        .order_by(StudentClassHistory.class_name)
        .all()
    )

    # Group students by their class name
    students_classes = defaultdict(list)
    for student, class_name in students:
        students_classes[class_name].append(student)

    # Instantiate forms for approve, deactivate, and regenerate actions
    approve_form = ApproveForm(prefix="approve")
    deactivate_form = ApproveForm(prefix="deactivate")
    regenerate_form = ApproveForm(prefix="regenerate")
    
    return render_template(
        "admin/students/approve_students.html",
        students_classes=students_classes,
        approve_form=approve_form,
        deactivate_form=deactivate_form,
        regenerate_form=regenerate_form,
        current_session=current_session,
        current_term=current_term,
    )


@admin_bp.route("/admin/approve_student/<string:class_name>/<int:student_id>", methods=["POST"])
@login_required
def approve_student(class_name, student_id):
    """
    Approves a student with the given student_id.

    Args:
        student_id (int): The ID of the student to be approved.

    Returns:
        redirect: Redirects to the "approve_students" route.

    Raises:
        403: If the current user is not an admin.

    """
    if not current_user.is_admin:
        abort(403)  # Forbidden access

    form = ApproveForm(prefix="approve")
    class_name = unquote(class_name).strip()

    if form.validate_on_submit():
        student = Student.query.get_or_404(student_id)
        if student.approved:
            flash(
                f"Student {student.first_name} {student.last_name} is already approved.",
                "alert alert-info",
            )
        else:
            student.approved = True
            db.session.commit()
            flash(
                f"Student {student.first_name} {student.last_name} has been approved.",
                "alert alert-success",
            )
    else:
        flash("Form validation failed. Please try again.", "alert alert-danger")

    return redirect(url_for("admins.students_by_class", class_name=class_name))


@admin_bp.route("/admin/deactivate_student/<string:class_name>/<int:student_id>", methods=["POST"])
@login_required
def deactivate_student(class_name, student_id):
    if not current_user.is_admin:
        abort(403)  # Forbidden access

    form = ApproveForm()
    class_name = unquote(class_name).strip()
    if form.validate_on_submit():
        student = Student.query.get_or_404(student_id)
        student.approved = False
        db.session.commit()
        flash(
            f"Student {student.first_name} {student.last_name} has been deactivated.",
            "alert alert-success",
        )
    else:
        flash("An error occurred. Please try again.", "alert alert-danger")
    return redirect(url_for("admins.students_by_class", class_name=class_name))


@admin_bp.route("/admin/regenerate_credentials/<string:class_name>/<int:student_id>", methods=["POST"])
@login_required
def regenerate_password(class_name, student_id):
    """
    Regenerates the password for a student's account and updates it in the database.

    Args:
        student_id (int): The ID of the student whose password needs to be regenerated.

    Returns:
        redirect: Redirects to the "approve_students" route after regenerating the password.

    Raises:
        403: If the current user is not an admin, a Forbidden access error is raised.

    """
    if not current_user.is_admin:
        abort(403)  # Forbidden access

    form = ApproveForm(prefix="regenerate")
    class_name = unquote(class_name).strip()

    if form.validate_on_submit():
        student = Student.query.get_or_404(student_id)
        new_password = "".join(
            random.choices(string.ascii_letters + string.digits, k=8)
        )
        student.user.set_password(new_password)
        db.session.commit()

        flash(
            f"Credentials regenerated for {student.first_name} {student.last_name}. New password: {new_password}",
            "alert alert-success",
        )
    else:
        flash("Form validation failed. Please try again.", "alert alert-danger")

    return redirect(url_for("admins.students_by_class", class_name=class_name))


@admin_bp.route("/select_class", methods=["GET", "POST"])
@login_required
def select_class():
    class_form = classForm()

    if class_form.validate_on_submit():
        class_name = class_form.class_name.data.strip()  # Ensure clean input
        
        # Redirect to the students_by_class route with class_name as a URL path segment
        return redirect(
            url_for(
                "admins.students_by_class",
                class_name=class_name
            )
        )
        
    return render_template(
        "admin/classes/select_class.html",
        class_form=class_form,
    )



@admin_bp.route("/students_by_class/<string:class_name>", methods=["GET", "POST"])
@login_required
def students_by_class(class_name):
    """
    Retrieve students by class for the current session.

    This function retrieves students in a specific class for the current session and term.
    It also handles the delete functionality.

    Returns:
        render_template: The rendered template with the students, session, class name, and form.
    """
    
    # # Decode and normalize class_name
    class_name = unquote(class_name).strip()
    
    if not class_name:
        flash("Invalid class name provided.", "alert alert-danger")
        return redirect(url_for("admins.select_class"))

    # Get current session and term directly from the Session model
    current_session, current_term = Session.get_current_session_and_term(include_term=True)

    # Initialize the form for deletion
    form = DeleteForm()

    # Retrieve the session ID based on the current session
    selected_session_id = Session.get_current_session().id

    # Fetch the student history filtered by session and class
    student_histories = StudentClassHistory.query.filter_by(
        session_id=selected_session_id, class_name=class_name
    ).all()

    # Extract students based on the history records
    students = [history.student for history in student_histories]

    if not students:
        flash(
            f"No students found in {class_name} for {current_session} - {current_term}",
            "alert alert-danger",
        )
        return redirect(url_for("admins.select_class"))

    return render_template(
        "admin/classes/students_by_class.html",
        students=students,
        session=current_session,
        current_session=current_session,
        class_name=class_name,
        form=form,
        current_term=current_term,
    )



@admin_bp.route("/toggle_fee_status/<string:class_name>/<int:student_id>", methods=["POST"])
@login_required
def toggle_fee_status(class_name, student_id):
    """
    Toggles the payment status for accessing results online for a student.
    """
    if not current_user.is_admin:
        abort(403)

    student = Student.query.get_or_404(student_id)
    class_name = unquote(class_name).strip()
    student.has_paid_fee = not student.has_paid_fee
    db.session.commit()

    flash(
        f"Payment status for {student.first_name} {student.last_name} has been updated.",
        "alert alert-success",
    )
    return redirect(url_for("admins.students_by_class", class_name=class_name))


# @admin_bp.route("/promote_student/<int:student_id>", methods=["POST"])
# @login_required
# def promote_student(student_id):
#     """
#     Promotes a student to the next class if applicable.

#     Args:
#         student_id (int): The ID of the student to be promoted.

#     Returns:
#         redirect: Redirects to the page displaying students of the new class.

#     Raises:
#         403: If the current user is not an admin.
#         404: If the student with the given ID is not found.
#     """
#     if not current_user.is_admin:
#         abort(403)  # Restrict access for non-admins

#     student = Student.query.get_or_404(student_id)
#     current_session = Session.get_current_session()

#     if not current_session:
#         flash("No current session available for promotion.", "alert alert-danger")
#         return redirect(url_for("admins.manage_students"))

#     # Define the class hierarchy
#     class_hierarchy = [
#         "Creche",
#         "Pre-Nursery",
#         "Nursery 1",
#         "Nursery 2",
#         "Nursery 3",
#         "Basic 1",
#         "Basic 2",
#         "Basic 3",
#         "Basic 4",
#         "Basic 5",
#         "Basic 6",
#         "JSS 1",
#         "JSS 2",
#         "JSS 3",
#     ]

#     # Retrieve the latest class from StudentClassHistory
#     latest_class_history = (
#         StudentClassHistory.query.filter_by(student_id=student.id)
#         .order_by(StudentClassHistory.id.desc())
#         .first()
#     )

#     if not latest_class_history:
#         flash("No class history found for the student.", "alert alert-danger")
#         return redirect(url_for("admins.manage_students"))

#     current_class = latest_class_history.class_name

#     # Promote the student to the next class if applicable
#     if current_class in class_hierarchy:
#         current_index = class_hierarchy.index(current_class)
#         if current_index + 1 < len(class_hierarchy):
#             new_class = class_hierarchy[current_index + 1]
#         else:
#             flash(
#                 "This student has completed the highest class.", "alert alert-warning"
#             )
#             return redirect(
#                 url_for(
#                     "admins.students_by_class",
#                     session_id=current_session.id,
#                     class_name=current_class,
#                 )
#             )
#     else:
#         flash("Current class not found in the hierarchy.", "alert alert-danger")
#         return redirect(
#             url_for(
#                 "admins.students_by_class",
#                 session_id=current_session.id,
#                 class_name=current_class,
#             )
#         )

#     # Add a new StudentClassHistory record for the promotion, only for the current session
#     new_class_history = StudentClassHistory(
#         student_id=student.id, session_id=current_session.id, class_name=new_class
#     )
#     db.session.add(new_class_history)
#     db.session.commit()

#     flash(
#         f"{student.first_name} has been promoted to {new_class}.", "alert alert-success"
#     )
#     return redirect(
#         url_for(
#             "admins.students_by_class",
#             session_id=current_session.id,
#             class_name=new_class,
#         )
#     )


# @admin_bp.route("/promote_student/<int:student_id>", methods=["POST"])
# @login_required
# def promote_student(student_id):
#     """
#     Promotes a student to the next class if applicable and ensures only one class record 
#     exists for the student in the current session.

#     Args:
#         student_id (int): The ID of the student to be promoted.

#     Returns:
#         redirect: Redirects to the page displaying students of the new class.

#     Raises:
#         403: If the current user is not an admin.
#         404: If the student with the given ID is not found.
#     """
#     if not current_user.is_admin:
#         abort(403)  # Restrict access for non-admins

#     student = Student.query.get_or_404(student_id)
#     current_session = Session.get_current_session()

#     if not current_session:
#         flash("No current session available for promotion.", "alert alert-danger")
#         return redirect(url_for("admins.manage_students"))

#     # Define the class hierarchy
#     class_hierarchy = [
#         "Creche",
#         "Pre-Nursery",
#         "Nursery 1",
#         "Nursery 2",
#         "Nursery 3",
#         "Basic 1",
#         "Basic 2",
#         "Basic 3",
#         "Basic 4",
#         "Basic 5",
#         "Basic 6",
#         "JSS 1",
#         "JSS 2",
#         "JSS 3",
#     ]

#     # Retrieve the latest class from StudentClassHistory
#     latest_class_history = (
#         StudentClassHistory.query.filter_by(student_id=student.id)
#         .order_by(StudentClassHistory.id.desc())
#         .first()
#     )

#     if not latest_class_history:
#         flash("No class history found for the student.", "alert alert-danger")
#         return redirect(url_for("admins.manage_students"))

#     current_class = latest_class_history.class_name

#     # Promote the student to the next class if applicable
#     if current_class in class_hierarchy:
#         current_index = class_hierarchy.index(current_class)
#         if current_index + 1 < len(class_hierarchy):
#             new_class = class_hierarchy[current_index + 1]
#         else:
#             flash(
#                 "This student has completed the highest class.", "alert alert-warning"
#             )
#             return redirect(
#                 url_for(
#                     "admins.students_by_class",
#                     class_name=current_class,
#                 )
#             )
#     else:
#         flash("Current class not found in the hierarchy.", "alert alert-danger")
#         return redirect(
#             url_for(
#                 "admins.students_by_class",
#                 class_name=current_class,
#             )
#         )

#     # Check for any existing record for this student in the current session
#     existing_record = StudentClassHistory.query.filter_by(
#         student_id=student.id, session_id=current_session.id
#     ).first()

#     # Delete the existing record if found
#     if existing_record:
#         db.session.delete(existing_record)

#     # Add the new StudentClassHistory record for the promotion
#     new_class_history = StudentClassHistory(
#         student_id=student.id, session_id=current_session.id, class_name=new_class
#     )
#     db.session.add(new_class_history)
#     db.session.commit()

#     flash(
#         f"{student.first_name} has been promoted to {new_class}.", "alert alert-success"
#     )
#     return redirect(
#         url_for(
#             "admins.students_by_class",
#             class_name=new_class,
#         )
#     )

@admin_bp.route("/promote_student/<string:class_name>/<int:student_id>", methods=["POST"])
@login_required
def promote_student(class_name, student_id):
    """
    Promotes a student to the next class in the next academic session.
    Ensures only one class record exists per session.
    """
    if not current_user.is_admin:
        abort(403)  # Restrict access for non-admins

    student = Student.query.get_or_404(student_id)
    current_session = Session.get_current_session()
    class_name = unquote(class_name).strip()

    if not current_session:
        flash("No current session available for promotion.", "alert alert-danger")
        return redirect(url_for("admins.students_by_class", class_name=class_name))

    # **Fix: Retrieve the next academic session**
    next_session = Session.query.filter(
        Session.year > current_session.year
    ).order_by(Session.year.asc()).first()

    if not next_session:
        flash("No next academic session available for promotion.", "alert alert-danger")
        return redirect(url_for("admins.students_by_class", class_name=class_name))

    # Define the class hierarchy
    class_hierarchy = [
        "Creche",
        "Pre-Nursery",
        "Nursery 1",
        "Nursery 2",
        "Nursery 3",
        "Basic 1",
        "Basic 2",
        "Basic 3",
        "Basic 4",
        "Basic 5",
        "Basic 6",
        "JSS 1",
        "JSS 2",
        "JSS 3",
    ]

    # Retrieve the latest class from StudentClassHistory
    # latest_class_history = (
    #     StudentClassHistory.query.filter_by(student_id=student.id)
    #     .order_by(StudentClassHistory.id.desc())
    #     .first()
    # )

    # if not latest_class_history:
    #     flash("No class history found for the student.", "alert alert-danger")
    #     return redirect(url_for("admins.students_by_class", class_name=class_name))

    current_class = class_name

    # Promote the student to the next class if applicable
    if current_class in class_hierarchy:
        current_index = class_hierarchy.index(current_class)
        if current_index + 1 < len(class_hierarchy):
            new_class = class_hierarchy[current_index + 1]
        else:
            flash("This student has completed the highest class.", "alert alert-warning")
            return redirect(
                url_for(
                    "admins.students_by_class",
                    class_name=current_class,
                )
            )
    else:
        flash("Current class not found in the hierarchy.", "alert alert-danger")
        return redirect(
            url_for(
                "admins.students_by_class",
                class_name=current_class,
            )
        )

    # **Fix: Ensure only one record exists for the next session**
    existing_record = StudentClassHistory.query.filter_by(
        student_id=student.id, session_id=next_session.id
    ).first()

    if existing_record:
        db.session.delete(existing_record)

    # Add a new StudentClassHistory record for the next session
    new_class_history = StudentClassHistory(
        student_id=student.id, session_id=next_session.id, class_name=new_class
    )
    db.session.add(new_class_history)
    db.session.commit()

    flash(
        f"{student.first_name} has been promoted to {new_class} in {next_session.year}.",
        "alert alert-success",
    )
    return redirect(
        url_for(
            "admins.students_by_class",
            class_name=new_class,
        )
    )




@admin_bp.route("/demote_student/<string:class_name>/<int:student_id>", methods=["POST"])
@login_required
def demote_student(class_name, student_id):
    """
    Demotes a student to the previous class in the next academic session.
    Ensures only one class record exists per session.
    """
    if not current_user.is_admin:
        abort(403)  # Restrict access for non-admins

    student = Student.query.get_or_404(student_id)
    current_session = Session.get_current_session()
    class_name = unquote(class_name).strip()

    if not current_session:
        flash("No current session available for demotion.", "alert alert-danger")
        return redirect(url_for("admins.students_by_class", class_name=class_name))

    # **Fix: Retrieve the next academic session**
    next_session = Session.query.filter(
        Session.year > current_session.year
    ).order_by(Session.year.asc()).first()

    if not next_session:
        flash("No next academic session available for demotion.", "alert alert-danger")
        return redirect(url_for("admins.students_by_class", class_name=class_name))

    # Define the class hierarchy
    class_hierarchy = [
        "Creche",
        "Pre-Nursery",
        "Nursery 1",
        "Nursery 2",
        "Nursery 3",
        "Basic 1",
        "Basic 2",
        "Basic 3",
        "Basic 4",
        "Basic 5",
        "Basic 6",
        "JSS 1",
        "JSS 2",
        "JSS 3",
    ]

    # # Retrieve the latest class from StudentClassHistory
    # latest_class_history = (
    #     StudentClassHistory.query.filter_by(student_id=student.id)
    #     .order_by(StudentClassHistory.id.desc())
    #     .first()
    # )

    # if not latest_class_history:
    #     flash("No class history found for the student.", "alert alert-danger")
    #     return redirect(url_for("admins.students_by_class", class_name=class_name))

    current_class = class_name

    # Demote the student to the previous class if applicable
    if current_class in class_hierarchy:
        current_index = class_hierarchy.index(current_class)
        if current_index > 0:
            new_class = class_hierarchy[current_index - 1]
        else:
            flash("This student is already in the lowest class.", "alert alert-warning")
            return redirect(
                url_for(
                    "admins.students_by_class",
                    class_name=current_class,
                )
            )
    else:
        flash("Current class not found in the hierarchy.", "alert alert-danger")
        return redirect(
            url_for(
                "admins.students_by_class",
                class_name=current_class,
            )
        )

    # **Fix: Ensure only one record exists for the next session**
    existing_record = StudentClassHistory.query.filter_by(
        student_id=student.id, session_id=next_session.id
    ).first()

    if existing_record:
        db.session.delete(existing_record)

    # Add a new StudentClassHistory record for the next session
    new_class_history = StudentClassHistory(
        student_id=student.id, session_id=next_session.id, class_name=new_class
    )
    db.session.add(new_class_history)
    db.session.commit()

    flash(
        f"{student.first_name} has been demoted to {new_class} in {next_session.year}.",
        "alert alert-success",
    )
    return redirect(
        url_for(
            "admins.students_by_class",
            class_name=new_class,
        )
    )




@admin_bp.route("/admin/edit_student/<int:student_id>", methods=["GET", "POST"])
@login_required
def edit_student(student_id):
    """
    Edit a student's information.

    Args:
        student_id (int): The ID of the student to be edited.

    Returns:
        A rendered template for editing a student's information.

    Raises:
        403: If the current user is not an admin.

    """
    if not current_user.is_admin:
        abort(403)  # Restrict access for non-admins

    student = Student.query.get_or_404(student_id)
    current_session = Session.get_current_session()

    if not current_session:
        flash("No current session found.", "alert alert-danger")
        return redirect(url_for("admins.students_by_class"))

    session_id = current_session.id
    form = EditStudentForm()

    # Retrieve the latest class history for the student in the current session
    student_class_history = (
        StudentClassHistory.query.filter_by(
            student_id=student.id, session_id=session_id
        )
        .order_by(StudentClassHistory.id.desc())
        .first()
    )

    if form.validate_on_submit():
        # Update student fields
        student.username = form.username.data
        student.first_name = form.first_name.data
        student.last_name = form.last_name.data
        student.middle_name = form.middle_name.data
        student.gender = form.gender.data

        # Update the class history for the current session
        if student_class_history:
            student_class_history.class_name = form.class_name.data
        else:
            # Create a new entry if no history exists for the current session
            new_class_history = StudentClassHistory(
                student_id=student.id,
                session_id=session_id,
                class_name=form.class_name.data,
            )
            db.session.add(new_class_history)

        # Update username in the User model
        user = User.query.filter_by(id=student.user_id).first()
        user.username = form.username.data

        db.session.commit()
        flash("Student updated successfully!", "alert alert-success")
        return redirect(
            url_for(
                "admins.students_by_class",
                session_id=session_id,
                class_name=form.class_name.data,
            )
        )

    elif request.method == "GET":
        # Pre-fill form with student data
        form.username.data = student.username
        form.first_name.data = student.first_name
        form.last_name.data = student.last_name
        form.middle_name.data = student.middle_name
        form.gender.data = student.gender

        # Pre-fill the class name from StudentClassHistory if available
        if student_class_history:
            form.class_name.data = student_class_history.class_name

    return render_template(
        "admin/students/edit_student.html", form=form, student=student
    )


@admin_bp.route("/delete_student_record/<string:class_name>/<int:student_id>", methods=["POST"])
@login_required
def delete_student_record(class_name, student_id):
    """
    Deletes a student's class record for the current session.

    Args:
        student_id (int): The ID of the student whose record is to be deleted.

    Returns:
        redirect: Redirects to the manage students page or class-specific view.

    Raises:
        403: If the current user is not an admin.
        404: If the student or current session record is not found.
    """
    if not current_user.is_admin:
        abort(403)  # Restrict access for non-admins

    # Check if the student exists
    student = Student.query.get_or_404(student_id)
    
    class_name = unquote(class_name).strip()

    # Get the current session dynamically
    current_session = Session.get_current_session()
    if not current_session:
        flash("No current session available.", "alert alert-danger")
        return redirect(url_for("admins.students_by_class", class_name=class_name))

    # Find the student's class record for the current session
    class_record = StudentClassHistory.query.filter_by(
        student_id=student.id, session_id=current_session.id
    ).first()

    if not class_record:
        flash(
            f"No class record found for {student.first_name} in the current session.",
            "alert alert-danger",
        )
        return redirect(url_for("admins.students_by_class", class_name=class_name))

    # Delete the class record
    db.session.delete(class_record)
    db.session.commit()

    flash(
        f"Class record for {student.first_name} in session {current_session.year} has been deleted.",
        "alert alert-success",
    )
    return redirect(url_for("admins.students_by_class", class_name=class_name))



@admin_bp.route("/admin/delete_student/<int:student_id>", methods=["POST"])
@login_required
def delete_student(student_id):
    """
    Delete a student and associated records.

    Args:
        student_id (int): The ID of the student to be deleted.

    Returns:
        redirect: Redirects to the student's class page.

    Raises:
        403: If the current user is not an admin.
        404: If the student with the given ID is not found.

    """
    if not current_user.is_admin:
        abort(403)  # Restrict access for non-admins

    form = DeleteForm()
    session_id = Session.query.filter_by(is_current=True).first().id

    if form.validate_on_submit():
        student = Student.query.get_or_404(student_id)
        student_class_history = (
            StudentClassHistory.query.filter_by(student_id=student.id)
            .order_by(StudentClassHistory.id.desc())
            .first()
        )

        try:
            # Delete all related results
            results = Result.query.filter_by(student_id=student.id).all()
            for result in results:
                db.session.delete(result)

            # Delete the associated User account
            user = User.query.get(student.user_id)
            if user:
                db.session.delete(user)

            # Delete all class history for the student
            class_history_records = StudentClassHistory.query.filter_by(
                student_id=student.id
            ).all()
            for history in class_history_records:
                db.session.delete(history)

            # Delete the student record itself
            db.session.delete(student)
            db.session.commit()

            flash(
                "Student and associated records deleted successfully!",
                "alert alert-success",
            )
        except Exception as e:
            db.session.rollback()
            flash(f"Error deleting student: {e}", "alert alert-danger")

    # Redirect to the student's class page, using the latest class from history or default if not found
    class_name = (
        student_class_history.class_name if student_class_history else "Unassigned"
    )
    return redirect(
        url_for(
            "admins.students_by_class", session_id=session_id, class_name=class_name
        )
    )


""""
Subject Management Section

This section displays the basic view of the subject management page
The page displays the list of subjects and their details
It also provides links to edit, delete, and add subjects

    manage_subjects: Displays the list of subjects and their details
    edit_subject: Edits the details of a subject
    delete_subject: Deletes a subject

"""


@admin_bp.route("/admin/manage_subjects", methods=["GET", "POST"])
@login_required
def manage_subjects():
    """
    Route handler for managing subjects in the admin panel.
    Allows adding and deactivating subjects for the current session.
    """
    if not current_user.is_authenticated or not current_user.is_admin:
        return redirect(url_for("auth.login"))

    form = SubjectForm()
    if form.validate_on_submit():
        subjects_input = form.name.data
        subject_names = [name.strip() for name in subjects_input.split(",")]

        for subject_name in subject_names:
            for section in form.section.data:
                existing_subject = Subject.query.filter_by(
                    name=subject_name, section=section, deactivated=False
                ).first()
                if existing_subject is None:
                    subject = Subject(name=subject_name, section=section)
                    db.session.add(subject)
        db.session.commit()
        flash("Subject(s) added successfully!", "alert alert-success")
        return redirect(url_for("admins.manage_subjects"))

    # Check for deactivation request
    if request.method == "POST" and "deactivate_subject_id" in request.form:
        subject_id = int(request.form.get("deactivate_subject_id"))
        subject = Subject.query.get(subject_id)
        if subject:
            subject.deactivated = True
            db.session.commit()
            flash(
                f"Subject '{subject.name}' has been deactivated.", "alert alert-warning"
            )
        return redirect(url_for("admins.manage_subjects"))

    # Display only active subjects for the current session
    subjects = (
        Subject.query.filter_by(deactivated=False).order_by(Subject.section).all()
    )
    subjects_by_section = {}
    for subject in subjects:
        if subject.section not in subjects_by_section:
            subjects_by_section[subject.section] = []
        subjects_by_section[subject.section].append(subject)

    delete_form = DeleteForm()
    return render_template(
        "admin/subjects/subject_admin.html",
        form=form,
        subjects_by_section=subjects_by_section,
        delete_form=delete_form,
    )


@admin_bp.route("/admin/edit_subject/<int:subject_id>", methods=["GET", "POST"])
@login_required
def edit_subject(subject_id):
    """
    Edit a subject with the given subject_id.

    Args:
        subject_id (int): The ID of the subject to be edited.

    Returns:
        A response object or a template with the form and subject data.

    Raises:
        404: If the subject with the given subject_id does not exist.
    """
    subject = Subject.query.get_or_404(subject_id)
    form = SubjectForm(obj=subject)
    if form.validate_on_submit():
        subject.name = form.name.data
        subject.section = form.section.data
        db.session.commit()

        # Update all related results with the new subject name and section
        results = Result.query.filter_by(subject_id=subject.id).all()
        for result in results:
            result.subject_name = subject.name
            result.subject_section = subject.section
        db.session.commit()

        flash("Subject updated successfully!", "alert alert-success")
        return redirect(url_for("admins.manage_subjects"))

    return render_template(
        "admin/subjects/edit_subject.html", form=form, subject=subject
    )


@admin_bp.route("/admin/delete_subject/<int:subject_id>", methods=["POST"])
def delete_subject(subject_id):
    """
    Delete a subject and its associated scores.

    Args:
        subject_id (int): The ID of the subject to be deleted.

    Returns:
        redirect: Redirects to the manage_subjects route.

    Raises:
        None

    """
    if not current_user.is_authenticated or not current_user.is_admin:
        return redirect(url_for("auth.login"))

    form = DeleteForm()  # Instantiate the DeleteForm

    if form.validate_on_submit():
        try:
            # Find the subject
            subject = Subject.query.get_or_404(subject_id)

            # Delete all scores associated with the subject
            Result.query.filter_by(subject_id=subject_id).delete()

            # Delete the subject
            db.session.delete(subject)
            db.session.commit()

            flash(
                "Subject and associated scores deleted successfully!",
                "alert alert-success",
            )
        except Exception as e:
            db.session.rollback()
            flash(f"Error deleting subject: {e}", "alert alert-danger")

    return redirect(url_for("admins.manage_subjects"))


""""
Result Management Section

Manages results by classes and sessions
This section allows the admin to view, edit, and delete results
It also provides the ability to generate broadsheets and download results

    select_term_session: Select the term and session for result management
    manage_results: Manage the results for a student
    broadsheet: Generate a broadsheet for a class
    download_broadsheet: Download a broadsheet for a class
    update_results: Update the results for a student
    calculate_results: Calculate the results for a student
    get_subjects_by_class_name: Get the subjects for a class
"""


@admin_bp.route("/manage_results/<int:student_id>", methods=["GET", "POST"])
@login_required
def manage_results(student_id):
    student = Student.query.get_or_404(student_id)

    # Get the current session and term directly from the Session model
    current_session, current_term = Session.get_current_session_and_term(include_term=True)

    if not current_session or not current_term:
        flash("Current session or term is not set", "alert alert-danger")
        return redirect(url_for("admins.students_by_class"))  # Updated redirection

    # Fetch student class history for the current session
    student_class = StudentClassHistory.get_class_by_session(
        student_id=student.id, session_year_str=current_session
    )

    if not student_class:
        flash(f"No class history for session {current_session}", "alert alert-danger")
        return redirect(url_for("admins.students_by_class"))  # Updated redirection

    # Get subjects based on class and session (with deactivated subjects based on session year)
    subjects = get_subjects_by_class_name(
        student_class, include_deactivated=(current_session == "2023/2024")
    )

    # Initialize the forms
    result_form = ResultForm(term=current_term, session=current_session)
    form = ManageResultsForm()

    # Fetch results in a single query and map by subject_id
    results = Result.query.filter_by(
        student_id=student.id, term=current_term, session=current_session
    ).all()
    results_dict = {result.subject_id: result for result in results}

    # If GET request, populate the form with existing results
    if request.method == "GET":
        populate_form_with_results(form, subjects, results_dict)

    # Handle POST request: update results if form is valid
    if form.validate_on_submit():
        update_results(student, current_term, current_session, form, result_form)
        flash("Results updated successfully", "alert alert-success")
        return redirect(
            url_for(
                "admins.manage_results",
                student_id=student.id,
                term=current_term,
                session=current_session,
            )
        )

    # Log validation errors if POST request fails
    if request.method == "POST":
        app.logger.error(f"Form validation failed: {form.errors}")

    # Calculate and display the results summary
    grand_total, average, cumulative_average, last_term_average = calculate_results(
        student.id, current_term, current_session
    )
    # Extract additional details from the first result (if available)
    next_term_begins = results[0].next_term_begins if results else None
    position = results[0].position if results else None
    date_issued = results[0].date_issued if results else None

    return render_template(
        "admin/results/manage_results.html",
        student=student,
        subjects=subjects,
        form=form,
        term=current_term,
        subject_results=zip(subjects, form.subjects),
        class_name=student_class,
        grand_total=grand_total,
        session=current_session,
        average=average,
        cumulative_average=cumulative_average,
        results_dict=results_dict,
        results=results,
        school_name="Aunty Anne's Int'l School",
        next_term_begins=next_term_begins,
        last_term_average=last_term_average,
        date_issued=date_issued,
        position=position,
    )


@admin_bp.route("/broadsheet/<string:class_name>", methods=["GET", "POST"])
@login_required
def broadsheet(class_name):
    try:
        # Decode and normalize class_name
        class_name = unquote(class_name).strip()
        
        # Get the current session and term directly from the Session model
        current_session, current_term = Session.get_current_session_and_term(include_term=True)

        if not current_session or not current_term:
            flash("Current session or term is not set", "alert-danger")
            return redirect(url_for("admins.students_by_class", class_name=class_name))

        # # Fetch the selected session object
        selected_session = Session.get_current_session()

        if not selected_session:
            flash(f"Session '{current_session}' not found.", "alert-danger")
            return redirect(url_for("admins.admin_dashboard"))

        # Fetch students and subjects based on class_name from StudentClassHistory
        student_class_histories = (
            StudentClassHistory.query.filter_by(
                class_name=class_name, session_id=selected_session.id
            )
            .options(joinedload(StudentClassHistory.student))
            .all()
        )

        students = [history.student for history in student_class_histories]
        subjects = get_subjects_by_class_name(class_name=class_name)

        if not students or not subjects:
            flash(
                f"No students or subjects found for class {class_name}.",
                "alert-info",
            )
            return render_template(
                "admin/results/broadsheet.html",
                students=[],
                subjects=[],
                broadsheet_data=[],
                subject_averages={},
                class_name=class_name,
            )

        broadsheet_data = []
        subject_averages = {
            subject.id: {"total": 0, "count": 0} for subject in subjects
        }

        # Iterate through students and their results
        for student in students:
            student_results = {
                "student": student,
                "results": {
                    subject.id: {
                        "class_assessment": "",
                        "summative_test": "",
                        "exam": "",
                        "total": "",
                        "grade": "",
                        "remark": "",
                    }
                    for subject in subjects
                },
                "grand_total": "",
                "average": "",
                "position": "",
            }

            # Fetch student results for the specific term and session
            results = Result.query.filter_by(
                student_id=student.id, term=current_term, session=current_session
            ).all()

            grand_total = 0
            non_zero_subjects = 0

            # Process each result for the student
            for result in results:
                if result.subject_id in student_results["results"]:
                    student_results["results"][result.subject_id] = {
                        "class_assessment": result.class_assessment or "",
                        "summative_test": result.summative_test or "",
                        "exam": result.exam or "",
                        "total": result.total or "",
                        "grade": result.grade or "",
                        "remark": result.remark or "",
                    }
                    if result.total and result.total > 0:
                        grand_total += result.total
                        non_zero_subjects += 1
                        subject_averages[result.subject_id]["total"] += result.total
                        subject_averages[result.subject_id]["count"] += 1
                    student_results["position"] = result.position or ""

            # Set grand total and average
            student_results["grand_total"] = grand_total if grand_total > 0 else ""
            average = (
                grand_total / non_zero_subjects if non_zero_subjects > 0 else ""
            )
            student_results["average"] = round(average, 1) if average else ""

            # Add student results to the broadsheet data
            broadsheet_data.append(student_results)

        # Calculate class averages for each subject
        for subject_id, values in subject_averages.items():
            values["average"] = (
                round(values["total"] / values["count"], 1)
                if values["count"]
                else ""
            )

        # Sort students by their average
        broadsheet_data.sort(key=lambda x: x["average"], reverse=True)

        return render_template(
            "admin/results/broadsheet.html",
            students=students,
            subjects=subjects,
            broadsheet_data=broadsheet_data,
            subject_averages=subject_averages,
            class_name=class_name,
            current_session=current_session,
            current_term=current_term,
        )

    except Exception as e:
        app.logger.error(
            f"Error generating broadsheet for class_name: {class_name} - {str(e)}"
        )
        flash("An error occurred. Please try again later.", "alert-danger")
        return redirect(url_for("admins.admin_dashboard"))


@admin_bp.route("/download_broadsheet/<string:class_name>", methods=["GET"])
def download_broadsheet(class_name):
    try:
        # Decode and normalize class_name
        class_name = unquote(class_name).strip()
        
        term = request.args.get("term")
        session = request.args.get("session")

        if not term or not session:
            flash(
                "Term and session must be provided to download broadsheet.",
                "alert-danger",
            )
            return redirect(url_for("admins.broadsheet", class_name=class_name))

        selected_session = Session.query.filter_by(year=session).first()
        if not selected_session:
            flash(f"Session '{session}' not found.", "alert-danger")
            return redirect(url_for("admins.admin_dashboard"))

        student_class_histories = (
            StudentClassHistory.query.filter_by(
                class_name=class_name, session_id=selected_session.id
            )
            .options(joinedload(StudentClassHistory.student))
            .all()
        )
        students = [history.student for history in student_class_histories]
        subjects = get_subjects_by_class_name(class_name=class_name)

        if not students or not subjects:
            flash(
                f"No students or subjects found for class {class_name}.", "alert-info"
            )
            return redirect(url_for("admins.broadsheet", class_name=class_name))

        broadsheet_data = []
        subject_averages = {
            subject.id: {"total": 0, "count": 0} for subject in subjects
        }

        for student in students:
            student_results = {
                "student": student,
                "results": {subject.id: None for subject in subjects},
                "grand_total": 0,
                "average": 0,
                "position": None,
            }
            results = Result.query.filter_by(
                student_id=student.id, term=term, session=session
            ).all()

            grand_total = 0
            non_zero_subjects = 0
            for result in results:
                student_results["results"][result.subject_id] = result
                if result.total > 0:
                    grand_total += result.total
                    non_zero_subjects += 1
                    subject_averages[result.subject_id]["total"] += result.total
                    subject_averages[result.subject_id]["count"] += 1
                student_results["position"] = result.position

            average = grand_total / non_zero_subjects if non_zero_subjects > 0 else 0
            student_results["grand_total"] = grand_total
            student_results["average"] = round(average, 1)
            broadsheet_data.append(student_results)

        for subject_id, values in subject_averages.items():
            values["average"] = (
                round(values["total"] / values["count"], 1)
                if values["count"] > 0
                else 0
            )

        broadsheet_data.sort(key=lambda x: x["average"], reverse=True)

        # Create Excel file
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Broadsheet_{class_name}_{term}"

        # Define styles
        header_font = Font(bold=True, size=16, name="Times New Roman")
        sub_header_font = Font(bold=True, size=12, name="Times New Roman")
        cell_font = Font(bold=True, size=12, name="Times New Roman")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        alignment = Alignment(horizontal="center", vertical="center")

        # Merge cells from A1 to the column for Class Average
        last_column = (
            2 + len(broadsheet_data) * 4
        )  # Adjust according to the number of students
        sheet.merge_cells(
            start_row=1, start_column=2, end_row=1, end_column=last_column
        )
        sheet["B1"].value = (
            f"Broadsheet for {class_name} - Term: {term}, Academic Session: {session}  -  Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        sheet["B1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["B1"].font = header_font

        # Set column widths
        subject_column_letter = get_column_letter(1)
        sheet.column_dimensions[subject_column_letter].width = (
            30  # Set subject column width to 30cm
        )

        class_average_column_letter = get_column_letter(last_column)
        sheet.column_dimensions[class_average_column_letter].width = (
            15  # Set class average width to 15 cm
        )

        # Write headers (now student names merged and centralized)
        headers = [""]
        for student_data in broadsheet_data:
            student = student_data["student"]
            headers.extend([f"{student.first_name} {student.last_name}", "", "", ""])
        headers.append("Class Average")
        sheet.append(headers)

        # Merge and center student names across their 4 corresponding columns
        for i, student_data in enumerate(
            broadsheet_data, start=2
        ):  # start=2 to account for header rows
            start_col = 2 + (i - 2) * 4
            end_col = start_col + 3

            # Set the width for C/A (start_col), S/T (start_col + 1), Exam (start_col + 2), and Total (start_col + 3)
            sheet.column_dimensions[get_column_letter(start_col)].width = 7  # C/A width
            sheet.column_dimensions[get_column_letter(start_col + 1)].width = (
                7  # S/T width
            )
            sheet.column_dimensions[get_column_letter(start_col + 2)].width = (
                7  # Exam width
            )
            sheet.column_dimensions[get_column_letter(start_col + 3)].width = (
                7  # Total width
            )

            sheet.merge_cells(
                start_row=2, start_column=start_col, end_row=2, end_column=end_col
            )
            cell = sheet.cell(row=2, column=start_col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = sub_header_font

        # Add sub-headers (C/A, S/T, Exam, Total) under each student name
        sub_headers = ["Subjects"]
        for _ in broadsheet_data:
            sub_headers.extend(["C/A", "S/T", "Exam", "Total"])
        sub_headers.append("")
        sheet.append(sub_headers)

        # Write data for each subject and student
        for subject in subjects:
            row = [subject.name]
            for student_data in broadsheet_data:
                result = student_data["results"][subject.id]
                if result:
                    row.extend(
                        [
                            result.class_assessment if result.class_assessment else "",
                            result.summative_test if result.summative_test else "",
                            result.exam if result.exam else "",
                            result.total if result.total else "",
                        ]
                    )
                else:
                    row.extend(["", "", "", ""])
            row.append(
                subject_averages[subject.id]["average"]
                if subject_averages[subject.id]["average"]
                else ""
            )
            sheet.append(row)

        # Write grand totals, averages, and positions, ensuring values are not empty
        sheet.append([""])
        sheet.append(
            ["Grand Total"]
            + sum(
                [
                    [
                        "",
                        "",
                        "",
                        (
                            student_data["grand_total"]
                            if student_data["grand_total"]
                            else ""
                        ),
                    ]
                    for student_data in broadsheet_data
                ],
                [],
            )
            + [""]
        )
        sheet.append(
            ["Average"]
            + sum(
                [
                    [
                        "",
                        "",
                        "",
                        student_data["average"] if student_data["average"] else "",
                    ]
                    for student_data in broadsheet_data
                ],
                [],
            )
            + [""]
        )
        sheet.append(
            ["Position"]
            + sum(
                [
                    ["", "", "", student_data["position"]]
                    for student_data in broadsheet_data
                ],
                [],
            )
            + [""]
        )

        # Set page orientation to landscape
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.PAPERSIZE_A4 = True

        for row in sheet.iter_rows(
            min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
        ):
            for cell in row:
                cell.border = border
                cell.font = cell_font

                # Apply left alignment specifically to "Subjects" column, sub-headers, and result cells
                if cell.column == 1:
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = alignment

        # Save to a bytes buffer
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        # Return the Excel file as a downloadable response
        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment;filename=Broadsheet_{class_name}_{term}_{session}.xlsx"
            },
        )

    except Exception as e:
        app.logger.error(
            f"Error generating broadsheet for class_name: {class_name} - {str(e)}"
        )
        flash(
            "An error occurred while downloading the broadsheet. Please try again later.",
            "alert-danger",
        )
        return redirect(url_for("admins.broadsheet", class_name=class_name))
